import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
import os
import re
warnings.filterwarnings('ignore')

# ============================================================================
# CONFIGURATION - SPECIFY YOUR FILE PATH HERE
# ============================================================================
#EXCEL_FILE_PATH = "C:/Users/begoo/Desktop/workingReportHistory-11019824.xlsx"
EXCEL_FILE_PATH = "C:/Users/begoo/Desktop/22ReportHistory-15177880.xlsx"
# ============================================================================

def calculate_pips(row):
    """Calculate pips based on entry/exit prices and trade type"""
    try:
        entry_price = row['Price']
        exit_price = row['Price.1']
        trade_type = row['Type']
        symbol = row['Symbol']
        
        if pd.isna(entry_price) or pd.isna(exit_price):
            return 0
        
        # Calculate raw price difference
        if trade_type == 'buy':
            price_diff = exit_price - entry_price
        else:  # sell
            price_diff = entry_price - exit_price
        
        # Determine pip multiplier based on symbol
        # For XAUUSD (gold), 1 pip = 0.01 (some brokers use 0.1)
        # For forex pairs, typically 1 pip = 0.0001 (for JPY pairs 0.01)
        symbol_upper = str(symbol).upper()
        
        if 'XAU' in symbol_upper or 'GOLD' in symbol_upper:
            # Gold - typically 0.01 or 0.1 depending on broker
            # Check the price format to determine
            if exit_price > 1000:  # Gold is typically quoted above 1000
                pip_value = 0.1  # Adjust if your broker uses 0.01
            else:
                pip_value = 0.01
        elif 'JPY' in symbol_upper:
            # JPY pairs use 0.01 as 1 pip
            pip_value = 0.01
        else:
            # Most other forex pairs use 0.0001 as 1 pip
            pip_value = 0.0001
        
        pips = price_diff / pip_value
        return round(pips, 1)
        
    except Exception as e:
        return 0

def extract_signal_number(comment):
    """Extract signal number from comment format: number_..."""
    if pd.isna(comment) or comment == '':
        return None
    
    # Look for pattern: number followed by underscore
    match = re.match(r'^(\d+)_', str(comment))
    if match:
        return match.group(1)
    return None

def analyze_signals(positions_df):
    """Analyze signals from comments"""
    signal_analysis = {}
    
    # Extract signal numbers from comments
    if 'Deal Comment' in positions_df.columns:
        positions_df['Signal'] = positions_df['Deal Comment'].apply(extract_signal_number)
    elif 'Comment' in positions_df.columns:
        positions_df['Signal'] = positions_df['Comment'].apply(extract_signal_number)
    else:
        return signal_analysis, positions_df
    
    # Filter only valid signals
    valid_signals = positions_df[positions_df['Signal'].notna()]
    
    if len(valid_signals) > 0:
        # Count unique signals
        signal_analysis['total_signals'] = valid_signals['Signal'].nunique()
        
        # Calculate legs per signal
        signal_counts = valid_signals['Signal'].value_counts()
        signal_analysis['avg_legs_per_signal'] = signal_counts.mean()
        signal_analysis['max_legs'] = signal_counts.max()
        signal_analysis['min_legs'] = signal_counts.min()
        
        # Calculate layers (legs / 4)
        signal_analysis['avg_layers'] = signal_analysis['avg_legs_per_signal'] / 4
        
        # Create detailed signal breakdown
        signal_details = []
        for signal in valid_signals['Signal'].unique():
            signal_data = valid_signals[valid_signals['Signal'] == signal]
            detail = {
                'Signal': signal,
                'Legs': len(signal_data),
                'Layers': len(signal_data) / 4,
                'Total P&L': signal_data['Profit'].sum() if 'Profit' in signal_data.columns else 0,
                'Total Pips': signal_data['Pips'].sum() if 'Pips' in signal_data.columns else 0,
                'Avg P&L': signal_data['Profit'].mean() if 'Profit' in signal_data.columns else 0,
                'Avg Pips': signal_data['Pips'].mean() if 'Pips' in signal_data.columns else 0,
            }
            signal_details.append(detail)
        
        signal_analysis['signal_breakdown'] = pd.DataFrame(signal_details)
    else:
        signal_analysis['total_signals'] = 0
        signal_analysis['avg_legs_per_signal'] = 0
        signal_analysis['avg_layers'] = 0
    
    return signal_analysis, positions_df

def read_mt5_report(filename):
    """Read the MT5 report and extract Positions and Deals tables"""
    
    # Read the entire Excel file to find where tables start
    df_all = pd.read_excel(filename, header=None)
    
    # Find where Positions and Deals sections start
    positions_idx = df_all[df_all[0] == 'Positions'].index[0] if 'Positions' in df_all[0].values else None
    deals_idx = df_all[df_all[0] == 'Deals'].index[0] if 'Deals' in df_all[0].values else None
    
    # Also find Orders section to know where Positions end
    orders_idx = None
    if 'Orders' in df_all[0].values:
        orders_idx = df_all[df_all[0] == 'Orders'].index[0]
    
    if positions_idx is None or deals_idx is None:
        raise ValueError("Could not find Positions or Deals sections in the file")
    
    # Determine where to stop reading Positions
    # Stop at Orders if it exists and comes before Deals, otherwise stop at Deals
    if orders_idx and orders_idx < deals_idx:
        positions_end = orders_idx
    else:
        positions_end = deals_idx
    
    # Read Positions table - stop before Orders or Deals section
    positions_df = pd.read_excel(filename, 
                                 skiprows=positions_idx + 1,  # Skip to header row
                                 nrows=positions_end - positions_idx - 2)  # Read until Orders/Deals section
    
    # Clean up positions_df - remove any rows that are completely empty or contain section headers
    positions_df = positions_df.dropna(how='all')  # Remove completely empty rows
    
    # Remove rows where the first column might contain section headers like "Orders"
    if not positions_df.empty and 'Time' in positions_df.columns:
        # Keep only rows where Time column has valid datetime-like data
        positions_df = positions_df[positions_df['Time'].notna()]
        positions_df = positions_df[~positions_df['Time'].astype(str).str.contains('Orders|Deals|Working', case=False, na=False)]
    
    # Read Deals table
    deals_df = pd.read_excel(filename, 
                             skiprows=deals_idx + 1)  # Skip to header row
    
    # Clean column names
    positions_df.columns = positions_df.columns.str.strip()
    deals_df.columns = deals_df.columns.str.strip()
    
    print(f"Positions data: {len(positions_df)} rows after cleaning")
    
    # Clean numeric columns that might have spaces or be formatted as strings
    numeric_cols_positions = ['Volume', 'Price', 'S / L', 'T / P', 'Price.1', 'Commission', 'Swap', 'Profit']
    for col in numeric_cols_positions:
        if col in positions_df.columns:
            # Convert to string first, remove spaces, then to numeric
            positions_df[col] = positions_df[col].astype(str).str.replace(' ', '').str.replace(',', '')
            positions_df[col] = pd.to_numeric(positions_df[col], errors='coerce')
    
    numeric_cols_deals = ['Volume', 'Price', 'Commission', 'Fee', 'Swap', 'Profit', 'Balance']
    for col in numeric_cols_deals:
        if col in deals_df.columns:
            # Convert to string first, remove spaces, then to numeric
            deals_df[col] = deals_df[col].astype(str).str.replace(' ', '').str.replace(',', '')
            deals_df[col] = pd.to_numeric(deals_df[col], errors='coerce')
    
    # Convert time columns to datetime
    for col in positions_df.columns:
        if 'Time' in col and positions_df[col].dtype == 'object':
            positions_df[col] = pd.to_datetime(positions_df[col], errors='coerce')
    
    for col in deals_df.columns:
        if 'Time' in col and deals_df[col].dtype == 'object':
            deals_df[col] = pd.to_datetime(deals_df[col], errors='coerce')
    
    # Add Pips calculation to positions_df
    positions_df['Pips'] = positions_df.apply(calculate_pips, axis=1)
    
    return positions_df, deals_df

def analyze_positions(positions_df):
    """Analyze positions data including pips and signals"""
    analysis = {}
    
    # Make a copy to avoid modifying the original
    positions_df = positions_df.copy()
    
    # Basic statistics
    analysis['total_positions'] = int(len(positions_df))
    analysis['unique_symbols'] = int(positions_df['Symbol'].nunique())
    analysis['symbols_list'] = positions_df['Symbol'].unique().tolist()
    
    # Position type breakdown
    type_counts = positions_df['Type'].value_counts()
    analysis['buy_positions'] = int(type_counts.get('buy', 0))
    analysis['sell_positions'] = int(type_counts.get('sell', 0))
    
    # Volume statistics - ensure numeric
    analysis['total_volume'] = float(positions_df['Volume'].sum())
    analysis['avg_volume'] = float(positions_df['Volume'].mean())
    analysis['max_volume'] = float(positions_df['Volume'].max())
    analysis['min_volume'] = float(positions_df['Volume'].min())
    
    # Pips analysis
    if 'Pips' in positions_df.columns:
        analysis['total_pips'] = float(positions_df['Pips'].sum())
        analysis['avg_pips'] = float(positions_df['Pips'].mean())
        analysis['max_pips_gain'] = float(positions_df['Pips'].max())
        analysis['max_pips_loss'] = float(positions_df['Pips'].min())
        analysis['winning_pips'] = float(positions_df[positions_df['Pips'] > 0]['Pips'].sum())
        analysis['losing_pips'] = float(positions_df[positions_df['Pips'] < 0]['Pips'].sum())
    
    # Profit/Loss analysis (if Profit column exists)
    if 'Profit' in positions_df.columns:
        profit_col = positions_df['Profit']
        analysis['total_pnl'] = float(profit_col.sum())
        analysis['avg_pnl'] = float(profit_col.mean())
        analysis['max_profit'] = float(profit_col.max())
        analysis['max_loss'] = float(profit_col.min())
        analysis['winning_positions'] = int((profit_col > 0).sum())
        analysis['losing_positions'] = int((profit_col < 0).sum())
        analysis['breakeven_positions'] = int((profit_col == 0).sum())
        analysis['win_rate'] = float((analysis['winning_positions'] / len(positions_df) * 100)) if len(positions_df) > 0 else 0.0
    
    # By symbol analysis
    symbol_stats = []
    for symbol in positions_df['Symbol'].unique():
        symbol_data = positions_df[positions_df['Symbol'] == symbol]
        symbol_stat = {
            'Symbol': symbol,
            'Count': int(len(symbol_data)),
            'Total Volume': float(symbol_data['Volume'].sum()),
            'Buy Count': int((symbol_data['Type'] == 'buy').sum()),
            'Sell Count': int((symbol_data['Type'] == 'sell').sum())
        }
        
        if 'Profit' in positions_df.columns:
            symbol_stat['Total P&L'] = float(symbol_data['Profit'].sum())
            symbol_stat['Win Rate %'] = float(((symbol_data['Profit'] > 0).sum() / len(symbol_data) * 100)) if len(symbol_data) > 0 else 0.0
        
        if 'Pips' in positions_df.columns:
            symbol_stat['Total Pips'] = float(symbol_data['Pips'].sum())
            symbol_stat['Avg Pips'] = float(symbol_data['Pips'].mean())
        
        symbol_stats.append(symbol_stat)
    
    analysis['symbol_breakdown'] = pd.DataFrame(symbol_stats)
    
    return analysis

def create_summary_workbook(input_filename, positions_df, deals_df, positions_analysis, signal_analysis):
    """Create a new workbook with summary and enhanced positions data"""
    
    # Create output filename in same directory as input
    input_dir = os.path.dirname(input_filename)
    output_filename = os.path.join(input_dir, "mt5_summary.xlsx")
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet if it exists
    for sheet in wb.sheetnames:
        wb.remove(wb[sheet])
    
    # Create Summary sheet first
    ws = wb.create_sheet('Summary', 0)
    
    # Define styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    subheader_font = Font(bold=True, size=11)
    subheader_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = 'MT5 TRADING REPORT SUMMARY'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].font = Font(italic=True)
    ws.merge_cells('A2:D2')
    
    current_row = 4
    
    # POSITIONS SUMMARY
    ws[f'A{current_row}'] = 'POSITIONS SUMMARY'
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws.merge_cells(f'A{current_row}:D{current_row}')
    current_row += 2
    
    # Positions metrics - INCLUDING PIPS AND SIGNALS
    positions_metrics = [
        ('Total Positions', positions_analysis['total_positions']),
        ('Unique Symbols', positions_analysis['unique_symbols']),
        ('Buy Positions', positions_analysis['buy_positions']),
        ('Sell Positions', positions_analysis['sell_positions']),
        ('Total Volume', positions_analysis['total_volume']),
        ('Average Volume', positions_analysis['avg_volume']),
    ]
    
    # Add signal metrics
    if 'total_signals' in signal_analysis:
        positions_metrics.extend([
            ('Total Signals', signal_analysis['total_signals']),
            ('Avg Legs per Signal', round(signal_analysis['avg_legs_per_signal'], 2) if signal_analysis['avg_legs_per_signal'] else 0),
            ('Avg Layers (Legs/4)', round(signal_analysis['avg_layers'], 2) if signal_analysis['avg_layers'] else 0),
        ])
    
    # Add pips metrics if available
    if 'total_pips' in positions_analysis:
        positions_metrics.extend([
            ('Total Pips', positions_analysis['total_pips']),
            ('Average Pips', positions_analysis['avg_pips']),
            ('Max Pips Gain', positions_analysis['max_pips_gain']),
            ('Max Pips Loss', positions_analysis['max_pips_loss']),
            ('Winning Pips', positions_analysis['winning_pips']),
            ('Losing Pips', positions_analysis['losing_pips']),
        ])
    
    if 'total_pnl' in positions_analysis:
        positions_metrics.extend([
            ('Total P&L', positions_analysis['total_pnl']),
            ('Average P&L', positions_analysis['avg_pnl']),
            ('Max Profit', positions_analysis['max_profit']),
            ('Max Loss', positions_analysis['max_loss']),
            ('Win Rate %', positions_analysis['win_rate']),
            ('Winning Positions', positions_analysis['winning_positions']),
            ('Losing Positions', positions_analysis['losing_positions']),
        ])
    
    for metric, value in positions_metrics:
        ws[f'A{current_row}'] = metric
        ws[f'B{current_row}'] = value  # Direct numeric value, no formatting
        ws[f'A{current_row}'].border = border
        ws[f'B{current_row}'].border = border
        
        # Apply number formatting based on metric type
        if 'Volume' in metric:
            ws[f'B{current_row}'].number_format = '#,##0.0000'
        elif 'Pips' in metric:
            ws[f'B{current_row}'].number_format = '#,##0.0'
        elif 'P&L' in metric or 'Profit' in metric or 'Loss' in metric:
            ws[f'B{current_row}'].number_format = '#,##0.00'
        elif 'Rate' in metric:
            ws[f'B{current_row}'].number_format = '0.00"%"'
        elif 'Layers' in metric or 'Legs' in metric:
            ws[f'B{current_row}'].number_format = '#,##0.00'
        elif isinstance(value, int):
            ws[f'B{current_row}'].number_format = '#,##0'
        elif isinstance(value, float):
            ws[f'B{current_row}'].number_format = '#,##0.00'
        
        current_row += 1
    
    current_row += 2
    
    # SYMBOL BREAKDOWN - NOW WITH PIPS
    if 'symbol_breakdown' in positions_analysis and not positions_analysis['symbol_breakdown'].empty:
        ws[f'A{current_row}'] = 'SYMBOL BREAKDOWN'
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].fill = header_fill
        ws.merge_cells(f'A{current_row}:I{current_row}')  # Extended for pips columns
        current_row += 2
        
        # Add symbol breakdown table
        symbol_df = positions_analysis['symbol_breakdown'].sort_values('Count', ascending=False)
        
        # Headers
        headers = list(symbol_df.columns)
        for col_idx, header in enumerate(headers):
            cell = ws.cell(row=current_row, column=col_idx+1, value=header)
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.border = border
        current_row += 1
        
        # Data - write raw numeric values
        for _, row in symbol_df.iterrows():
            for col_idx, value in enumerate(row):
                # Write the raw value directly
                cell = ws.cell(row=current_row, column=col_idx+1, value=value)
                cell.border = border
                
                # Apply formatting to numeric values
                if isinstance(value, (int, float)) and not pd.isna(value):
                    if 'P&L' in headers[col_idx] or 'Profit' in headers[col_idx]:
                        cell.number_format = '#,##0.00'
                    elif 'Pips' in headers[col_idx]:
                        cell.number_format = '#,##0.0'
                    elif 'Rate' in headers[col_idx]:
                        cell.number_format = '0.00"%"'
                    elif 'Volume' in headers[col_idx]:
                        cell.number_format = '#,##0.0000'
                    elif 'Count' in headers[col_idx]:
                        cell.number_format = '#,##0'
            current_row += 1
    
    # Adjust column widths - Fixed to handle merged cells
    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            try:
                # Skip merged cells
                if hasattr(cell, 'column_letter'):
                    if column_letter is None:
                        column_letter = cell.column_letter
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
            except:
                pass
        
        if column_letter and max_length > 0:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Now create the Positions with Comments sheet and get the enhanced dataframe
    positions_with_comments = create_positions_with_comments_sheet(wb, positions_df, deals_df)
    
    # Add Signal Analysis sheet if we have signal data
    if 'signal_breakdown' in signal_analysis and not signal_analysis['signal_breakdown'].empty:
        print("\nAdding Signal Analysis...")
        add_dataframe_to_sheet(wb, signal_analysis['signal_breakdown'], 'Signal Analysis', 2)
    
    # Add daily close analysis
    print("\nAdding Daily Close Summary...")
    daily_close_summary = analyze_daily_close_performance(positions_with_comments)
    if daily_close_summary is not None and not daily_close_summary.empty:
        add_dataframe_to_sheet(wb, daily_close_summary, 'Daily Close Summary', 3)
    
    # Add weekly summary
    print("Adding Weekly Summary...")
    weekly_summary = create_weekly_summary(positions_with_comments)
    if weekly_summary is not None and not weekly_summary.empty:
        add_dataframe_to_sheet(wb, weekly_summary, 'Weekly Summary', 4)
    
    # Save the workbook
    wb.save(output_filename)
    print(f"\nSummary workbook created successfully: {output_filename}")
    
    # Return both the filename and the positions with comments for further use
    return output_filename, wb, positions_with_comments

def create_positions_with_comments_sheet(wb, positions_df, deals_df):
    """Create a sheet with all positions data plus comments from deals and signal analysis"""
    
    # Create a copy of positions dataframe
    positions_with_comments = positions_df.copy()
    
    # Create a dictionary mapping Order numbers to Comments from deals
    order_to_comment = {}
    if 'Order' in deals_df.columns and 'Comment' in deals_df.columns:
        for _, row in deals_df.iterrows():
            order_num = str(row['Order']).strip()
            comment = row['Comment'] if pd.notna(row['Comment']) else ''
            if order_num and order_num != 'nan':
                order_to_comment[order_num] = comment
    
    # Add Comment column to positions data
    comments = []
    for _, row in positions_with_comments.iterrows():
        position_num = str(row['Position']).strip() if 'Position' in row else ''
        comment = order_to_comment.get(position_num, '')
        comments.append(comment)
    
    positions_with_comments['Deal Comment'] = comments
    
    # Add Signal column
    positions_with_comments['Signal'] = positions_with_comments['Deal Comment'].apply(extract_signal_number)
    
    # Calculate legs for each position's signal
    signal_counts = positions_with_comments[positions_with_comments['Signal'].notna()]['Signal'].value_counts()
    positions_with_comments['Signal Legs'] = positions_with_comments['Signal'].map(signal_counts).fillna(0).astype(int)
    positions_with_comments['Signal Layers'] = (positions_with_comments['Signal Legs'] / 4).round(2)
    
    # Create new sheet
    ws = wb.create_sheet('Positions with Comments', 1)
    
    # Define styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    headers = list(positions_with_comments.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Write data
    for row_idx, (_, row) in enumerate(positions_with_comments.iterrows(), 2):
        for col_idx, (col_name, value) in enumerate(row.items(), 1):
            # Handle datetime values
            if pd.api.types.is_datetime64_any_dtype(type(value)):
                value = value.strftime('%Y-%m-%d %H:%M:%S') if pd.notna(value) else ''
            # Handle NaN values
            elif pd.isna(value):
                value = ''
            
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            
            # Apply number formatting for numeric columns
            if col_name in ['Volume', 'Price', 'S / L', 'T / P', 'Price.1']:
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.0000'
            elif col_name == 'Pips':
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.0'
            elif col_name in ['Commission', 'Swap', 'Profit']:
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
            elif col_name == 'Position' or col_name == 'Signal Legs':
                if isinstance(value, (int, float)):
                    cell.number_format = '0'
            elif col_name == 'Signal Layers':
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            try:
                if hasattr(cell, 'column_letter'):
                    if column_letter is None:
                        column_letter = cell.column_letter
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
            except:
                pass
        
        if column_letter:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Apply autofilter to the data range
    if len(positions_with_comments) > 0:
        last_row = len(positions_with_comments) + 1
        last_col_letter = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    print(f"Positions with Comments sheet added with {len(positions_with_comments)} rows")
    
    # Return the enhanced dataframe for further analysis
    return positions_with_comments

def analyze_daily_close_performance(positions_df):
    """Analyze performance by close date (Time.1), including pips and signal analysis"""
    try:
        if 'Time.1' not in positions_df.columns or positions_df.empty:
            print("Time.1 column not found in positions data")
            return None
        
        # Create a copy for analysis
        df = positions_df.copy()
        
        # Convert Time.1 to datetime and extract date
        df['Close_DateTime'] = pd.to_datetime(df['Time.1'])
        df['Close_Date'] = df['Close_DateTime'].dt.date
        
        # Adjust Sunday trades to Monday
        df['Adjusted_Close_Date'] = df['Close_DateTime'].apply(
            lambda x: x.date() if x.weekday() != 6 else (x + pd.Timedelta(days=1)).date()
        )
        
        # Calculate trade duration in hours
        if 'Time' in df.columns:
            df['Open_DateTime'] = pd.to_datetime(df['Time'])
            df['Duration_Hours'] = (df['Close_DateTime'] - df['Open_DateTime']).dt.total_seconds() / 3600
        
        # Build aggregation dictionary
        agg_dict = {
            'Position': 'count',  # Number of closed trades
            'Profit': ['sum', 'mean', 'std', 'min', 'max'],  # P&L metrics
            'Volume': ['sum', 'mean'],  # Volume metrics
            'Symbol': lambda x: x.nunique(),  # Unique symbols
            'Type': lambda x: (x == 'buy').sum()  # Count of buy trades
        }
        
        # Add Pips aggregation if available
        if 'Pips' in df.columns:
            agg_dict['Pips'] = ['sum', 'mean', 'min', 'max']
        
        # Add Signal aggregation if available
        if 'Signal' in df.columns:
            agg_dict['Signal'] = lambda x: x.dropna().nunique()  # Count unique signals
        
        # Group by adjusted close date for daily statistics
        daily_stats = df.groupby('Adjusted_Close_Date').agg(agg_dict).round(2)
        
        # Flatten column names
        columns = [
            'Trades Closed', 
            'Total P&L', 
            'Avg P&L', 
            'P&L StdDev',
            'Min P&L', 
            'Max P&L', 
            'Total Volume',
            'Avg Volume',
            'Symbols Traded',
            'Buy Trades'
        ]
        
        # Add pips columns if they exist
        if 'Pips' in df.columns:
            columns.extend(['Total Pips', 'Avg Pips', 'Min Pips', 'Max Pips'])
        
        # Add signal column if it exists
        if 'Signal' in df.columns:
            columns.append('Unique Signals')
        
        daily_stats.columns = columns
        
        # Add duration if available
        if 'Duration_Hours' in df.columns:
            duration_stats = df.groupby('Adjusted_Close_Date')['Duration_Hours'].mean().round(2)
            daily_stats['Avg Duration (hrs)'] = duration_stats
        
        # Add Sell trades count
        daily_stats['Sell Trades'] = daily_stats['Trades Closed'] - daily_stats['Buy Trades']
        
        # Calculate win/loss statistics
        win_stats = df[df['Profit'] > 0].groupby('Adjusted_Close_Date').agg({
            'Position': 'count',
            'Profit': 'sum'
        })
        win_stats.columns = ['Winning Trades', 'Gross Profit']
        
        # Add winning pips if available
        if 'Pips' in df.columns:
            win_pips = df[df['Pips'] > 0].groupby('Adjusted_Close_Date')['Pips'].sum()
            win_stats['Winning Pips'] = win_pips
        
        loss_stats = df[df['Profit'] < 0].groupby('Adjusted_Close_Date').agg({
            'Position': 'count',
            'Profit': 'sum'
        })
        loss_stats.columns = ['Losing Trades', 'Gross Loss']
        
        # Add losing pips if available
        if 'Pips' in df.columns:
            loss_pips = df[df['Pips'] < 0].groupby('Adjusted_Close_Date')['Pips'].sum()
            loss_stats['Losing Pips'] = loss_pips
        
        # Merge win/loss stats
        daily_stats = daily_stats.merge(win_stats, left_index=True, right_index=True, how='left')
        daily_stats = daily_stats.merge(loss_stats, left_index=True, right_index=True, how='left')
        
        # Fill NaN values
        daily_stats['Winning Trades'] = daily_stats['Winning Trades'].fillna(0).astype(int)
        daily_stats['Losing Trades'] = daily_stats['Losing Trades'].fillna(0).astype(int)
        daily_stats['Gross Profit'] = daily_stats['Gross Profit'].fillna(0)
        daily_stats['Gross Loss'] = daily_stats['Gross Loss'].fillna(0)
        
        if 'Winning Pips' in daily_stats.columns:
            daily_stats['Winning Pips'] = daily_stats['Winning Pips'].fillna(0)
            daily_stats['Losing Pips'] = daily_stats['Losing Pips'].fillna(0)
        
        # Calculate additional metrics
        daily_stats['Win Rate %'] = (daily_stats['Winning Trades'] / daily_stats['Trades Closed'] * 100).round(2)
        daily_stats['Profit Factor'] = (daily_stats['Gross Profit'] / abs(daily_stats['Gross Loss'])).replace([np.inf, -np.inf], 0).round(2)
        
        # Add week information
        daily_stats = daily_stats.reset_index()
        daily_stats['Day of Week'] = pd.to_datetime(daily_stats['Adjusted_Close_Date']).dt.day_name()
        
        # Add week start date for grouping
        daily_stats['Week Starting'] = pd.to_datetime(daily_stats['Adjusted_Close_Date']).apply(
            lambda x: (x - pd.Timedelta(days=x.weekday())).date()
        )
        
        # Calculate cumulative P&L and Pips
        daily_stats['Cumulative P&L'] = daily_stats['Total P&L'].cumsum()
        if 'Total Pips' in daily_stats.columns:
            daily_stats['Cumulative Pips'] = daily_stats['Total Pips'].cumsum()
        
        # Reorder columns for better readability
        column_order = [
            'Adjusted_Close_Date', 'Day of Week', 'Week Starting',
            'Trades Closed', 'Winning Trades', 'Losing Trades', 'Win Rate %'
        ]
        
        # Add signal column if it exists
        if 'Unique Signals' in daily_stats.columns:
            column_order.append('Unique Signals')
        
        column_order.extend([
            'Total P&L', 'Cumulative P&L', 'Avg P&L', 'P&L StdDev',
            'Min P&L', 'Max P&L', 'Gross Profit', 'Gross Loss', 'Profit Factor'
        ])
        
        # Add pips columns to order if they exist
        if 'Total Pips' in daily_stats.columns:
            pips_order = ['Total Pips', 'Cumulative Pips', 'Avg Pips', 'Min Pips', 'Max Pips']
            if 'Winning Pips' in daily_stats.columns:
                pips_order.extend(['Winning Pips', 'Losing Pips'])
            column_order.extend(pips_order)
        
        column_order.extend(['Total Volume', 'Avg Volume', 'Buy Trades', 'Sell Trades', 'Symbols Traded'])
        
        # Add duration column if it exists
        if 'Avg Duration (hrs)' in daily_stats.columns:
            column_order.append('Avg Duration (hrs)')
        
        # Ensure all columns exist before reordering
        available_columns = [col for col in column_order if col in daily_stats.columns]
        daily_stats = daily_stats[available_columns]
        
        # Rename the date column for clarity
        daily_stats = daily_stats.rename(columns={'Adjusted_Close_Date': 'Close Date'})
        
        # Sort by date
        daily_stats = daily_stats.sort_values('Close Date')
        
        return daily_stats
        
    except Exception as e:
        print(f"Error in daily close analysis: {e}")
        import traceback
        traceback.print_exc()
        return None

def create_weekly_summary(positions_df):
    """Create weekly summary from the positions data including pips and signals"""
    try:
        if 'Time.1' not in positions_df.columns or positions_df.empty:
            return None
            
        # Create a copy for analysis
        df = positions_df.copy()
        
        # Convert Time.1 to datetime
        df['Close_DateTime'] = pd.to_datetime(df['Time.1'])
        
        # Adjust Sunday trades to Monday
        df['Adjusted_Close_Date'] = df['Close_DateTime'].apply(
            lambda x: x.date() if x.weekday() != 6 else (x + pd.Timedelta(days=1)).date()
        )
        
        # Group by week
        df['Week_Start'] = pd.to_datetime(df['Adjusted_Close_Date']).apply(
            lambda x: (x - pd.Timedelta(days=x.weekday())).date()
        )
        
        # Build aggregation dictionary
        agg_dict = {
            'Position': 'count',
            'Profit': ['sum', 'mean', 'min', 'max'],
            'Volume': 'sum',
            'Symbol': lambda x: x.nunique()
        }
        
        # Add Pips aggregation if available
        if 'Pips' in df.columns:
            agg_dict['Pips'] = ['sum', 'mean', 'min', 'max']
        
        # Add Signal aggregation if available
        if 'Signal' in df.columns:
            agg_dict['Signal'] = lambda x: x.dropna().nunique()
        
        weekly_stats = df.groupby('Week_Start').agg(agg_dict).round(2)
        
        # Build column names
        columns = [
            'Total Trades', 'Total P&L', 'Avg P&L', 'Min P&L', 'Max P&L',
            'Total Volume', 'Symbols Traded'
        ]
        
        # Add pips columns if they exist
        if 'Pips' in df.columns:
            columns.extend(['Total Pips', 'Avg Pips', 'Min Pips', 'Max Pips'])
        
        # Add signal column if it exists
        if 'Signal' in df.columns:
            columns.append('Unique Signals')
        
        weekly_stats.columns = columns
        
        # Add win rate
        win_counts = df[df['Profit'] > 0].groupby('Week_Start').size()
        total_counts = df.groupby('Week_Start').size()
        weekly_stats['Win Rate %'] = (win_counts / total_counts * 100).round(2)
        weekly_stats['Win Rate %'] = weekly_stats['Win Rate %'].fillna(0)
        
        # Calculate cumulative P&L and Pips
        weekly_stats['Cumulative P&L'] = weekly_stats['Total P&L'].cumsum()
        if 'Total Pips' in weekly_stats.columns:
            weekly_stats['Cumulative Pips'] = weekly_stats['Total Pips'].cumsum()
        
        # Reset index
        weekly_stats = weekly_stats.reset_index()
        weekly_stats['Week Ending'] = pd.to_datetime(weekly_stats['Week_Start']) + pd.Timedelta(days=6)
        
        # Reorder columns
        column_order = [
            'Week_Start', 'Week Ending', 'Total Trades'
        ]
        
        # Add signal column if it exists
        if 'Unique Signals' in weekly_stats.columns:
            column_order.append('Unique Signals')
        
        column_order.extend([
            'Total P&L', 'Cumulative P&L',
            'Avg P&L', 'Min P&L', 'Max P&L'
        ])
        
        # Add pips columns if they exist
        if 'Total Pips' in weekly_stats.columns:
            column_order.extend(['Total Pips', 'Cumulative Pips', 'Avg Pips', 'Min Pips', 'Max Pips'])
        
        column_order.extend(['Win Rate %', 'Total Volume', 'Symbols Traded'])
        
        # Ensure all columns exist before reordering
        available_columns = [col for col in column_order if col in weekly_stats.columns]
        weekly_stats = weekly_stats[available_columns]
        
        # Rename for clarity
        weekly_stats = weekly_stats.rename(columns={'Week_Start': 'Week Starting'})
        
        return weekly_stats
        
    except Exception as e:
        print(f"Error in weekly summary: {e}")
        return None

def add_dataframe_to_sheet(workbook, df, sheet_name, position):
    """Add a dataframe as a new sheet in the workbook"""
    
    # Create new sheet
    ws = workbook.create_sheet(sheet_name, position)
    
    # Define styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_idx, header in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=str(header))
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Write data
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, value in enumerate(row, 1):
            # Handle different data types
            if pd.isna(value):
                value = ''
            elif isinstance(value, pd.Timestamp):
                value = value.strftime('%Y-%m-%d %H:%M:%S')
            elif isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d')
            elif hasattr(value, 'date'):  # Handle date objects
                value = value.strftime('%Y-%m-%d') if hasattr(value, 'strftime') else str(value)
            
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            
            # Apply number formatting based on column name
            col_name = df.columns[col_idx-1]
            if 'P&L' in col_name or 'Profit' in col_name or 'Loss' in col_name:
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
            elif 'Pips' in col_name:
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.0'
            elif 'Volume' in col_name:
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.0000'
            elif 'Layers' in col_name or 'Legs' in col_name:
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
            elif 'Rate' in col_name or 'Factor' in col_name:
                if isinstance(value, (int, float)):
                    cell.number_format = '0.00'
            elif isinstance(value, int):
                cell.number_format = '#,##0'
            elif isinstance(value, float):
                cell.number_format = '#,##0.00'
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            try:
                if hasattr(cell, 'column_letter'):
                    if column_letter is None:
                        column_letter = cell.column_letter
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
            except:
                pass
        
        if column_letter:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Apply autofilter
    if len(df) > 0:
        last_row = len(df) + 1
        last_col_letter = get_column_letter(len(df.columns))
        ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
    
    # Freeze header row
    ws.freeze_panes = 'A2'

def main():
    # Use the file path specified at the top of the file
    filename = EXCEL_FILE_PATH
    
    # Check if file exists
    if not os.path.exists(filename):
        print(f"Error: File not found at {filename}")
        print("Please check the EXCEL_FILE_PATH variable at the top of this script.")
        return
    
    try:
        print(f"Reading MT5 report from: {filename}")
        positions_df, deals_df = read_mt5_report(filename)
        
        print(f"Found {len(positions_df)} positions and {len(deals_df)} deals")
        print(f"Calculated pips for all positions")
        
        print("\nAnalyzing positions...")
        positions_analysis = analyze_positions(positions_df)
        
        # Create positions with comments first to get signal data
        print("\nCreating positions with comments and analyzing signals...")
        
        # Need to create a temporary positions_with_comments for signal analysis
        order_to_comment = {}
        if 'Order' in deals_df.columns and 'Comment' in deals_df.columns:
            for _, row in deals_df.iterrows():
                order_num = str(row['Order']).strip()
                comment = row['Comment'] if pd.notna(row['Comment']) else ''
                if order_num and order_num != 'nan':
                    order_to_comment[order_num] = comment
        
        positions_df_temp = positions_df.copy()
        comments = []
        for _, row in positions_df_temp.iterrows():
            position_num = str(row['Position']).strip() if 'Position' in row else ''
            comment = order_to_comment.get(position_num, '')
            comments.append(comment)
        positions_df_temp['Deal Comment'] = comments
        
        # Analyze signals
        signal_analysis, positions_df_with_signals = analyze_signals(positions_df_temp)
        positions_df = positions_df_with_signals  # Update positions_df with signal column
        
        print(f"Found {signal_analysis.get('total_signals', 0)} unique signals")
        
        print("\nCreating summary workbook with analysis sheets...")
        output_file, workbook, positions_with_comments = create_summary_workbook(
            filename, positions_df, deals_df, positions_analysis, signal_analysis
        )
        
        # Keep positions_with_comments in memory for further analysis
        print(f"\nPositions with comments data loaded: {len(positions_with_comments)} rows")
        print(f"Columns available: {list(positions_with_comments.columns)}")
        
        # Return the dataframe so it can be used for further analysis
        return {
            'positions_with_comments': positions_with_comments,
            'positions_df': positions_df,
            'deals_df': deals_df,
            'positions_analysis': positions_analysis,
            'signal_analysis': signal_analysis,
            'output_file': output_file
        }
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    # Run the main analysis and get the data back
    analysis_data = main()
    
    # The data is now available for further analysis
    if analysis_data:
        print("\n" + "="*50)
        print("DATA AVAILABLE IN MEMORY FOR FURTHER ANALYSIS")
        print("="*50)
        print("Access the following variables:")
        print("  - analysis_data['positions_with_comments']: Full positions data with comments, pips, and signals")
        print("  - analysis_data['positions_df']: Original positions data with pips and signals")
        print("  - analysis_data['deals_df']: Deals data")
        print("  - analysis_data['positions_analysis']: Position statistics including pips")
        print("  - analysis_data['signal_analysis']: Signal statistics and breakdown")
        print("  - analysis_data['output_file']: Output file path")
        print("\nExcel file created with the following sheets:")
        print("  1. Summary - Overall trading statistics with PIPS and SIGNALS")
        print("  2. Positions with Comments - Full positions data with PIPS, SIGNALS, LEGS, LAYERS")
        print("  3. Signal Analysis - Detailed breakdown of each signal")
        print("  4. Daily Close Summary - Performance by close date with SIGNALS")
        print("  5. Weekly Summary - Weekly performance overview with SIGNALS")
        print("\nSignal analysis complete - Legs and Layers calculated!")